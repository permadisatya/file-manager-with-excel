#!/usr/bin/env python

from genericpath import exists
import os
import pathlib
from openpyxl import load_workbook, Workbook
from posixpath import join
import argparse

fPath = os.getcwd()
xlsxFile = os.path.join(fPath, "LOG.XLSX")
txtFile = os.path.join(fPath, "FOLDER.TXT")

# check files required is exist
def isExist(path):
    status = os.path.isfile(path)
    return status

if isExist(path=txtFile) == True:
    pass
else:
    open(txtFile, "x")

if isExist(path=xlsxFile) == True:
    pass
else:
    wb = Workbook()
    del wb['Sheet']
    wb.create_sheet("tbl_file", 0)
    wb.create_sheet("tbl_log", 1)
    wb["tbl_file"].cell(row=1, column=1).value = "id_files"
    wb["tbl_file"].cell(row=1, column=2).value = "path"
    wb["tbl_file"].cell(row=1, column=3).value = "file_name"
    wb["tbl_file"].cell(row=1, column=4).value = "status"
    wb["tbl_file"].cell(row=1, column=5).value = "new_file_name"
    wb["tbl_file"].cell(row=1, column=6).value = "rename_status"
    wb["tbl_log"].cell(row=1, column=1).value = "id_files"
    wb["tbl_log"].cell(row=1, column=2).value = "path"
    wb["tbl_log"].cell(row=1, column=3).value = "file_name"
    wb["tbl_log"].cell(row=1, column=4).value = "old_file_name"
    wb.save(xlsxFile)

# text to maintain
txt = open(txtFile).read().splitlines()
listFolder = []
for path in txt:
    listFolder.append(path.lstrip())

# xlsx to maintain
xlsx = load_workbook(xlsxFile)
tableFile = xlsx["tbl_file"]
tableLog = xlsx["tbl_log"]

# get all file information in certain folder
def listFile(path):
    a = []
    b = []
    for entry in pathlib.Path(path).iterdir():
        if entry.is_file():
            a.append("".join(["ID", str(os.path.getctime(entry)).replace(".", "")]))
            b.append(entry.name)
    return a, b

# get last non-empty row
def lastRow(col):
    row = []
    for i in col:
        if i.value != "" and i.value != None:
            row.append(i.row)
    return max(row)

# get data column
def listData(path):
    a = []
    b = []
    c = []
    for i in tableFile["A"]:
        if tableFile.cell(row=i.row, column=3).value == path:
            a.append(tableFile.cell(row=i.row, column=1).row)
            b.append(tableFile.cell(row=i.row, column=1).value)
            c.append(tableFile.cell(row=i.row, column=2).value)
    return a, b, c

# get data column
def listRename(str):
    a = []
    b = []
    c = []
    d = []
    e = []
    for i in tableFile["A"]:
        if tableFile.cell(row=i.row, column=6).value == str:
            a.append(tableFile.cell(row=i.row, column=1).row)
            b.append(tableFile.cell(row=i.row, column=1).value)
            c.append(tableFile.cell(row=i.row, column=2).value)
            d.append(tableFile.cell(row=i.row, column=3).value)
            e.append(tableFile.cell(row=i.row, column=5).value)
    return a, b, c, d, e

# update tableFile
def insertData(
    status, 
    row, 
    folderPath, 
    fileID = None, 
    fileName = None,
    newFileName = None, 
    newRow = None,
):
    tableFile.cell(row=row, column=3).value = folderPath
    tableFile.cell(row=row, column=4).value = status

    if status == "New":
        tableFile.cell(row=row, column=1).value = fileID
        tableFile.cell(row=row, column=2).value = fileName

    if status == "Renamed":
        tableFile.cell(row=row, column=1).value = fileID
        tableFile.cell(row=row, column=2).value = newFileName
        tableFile.cell(row=row, column=5).value = None
        tableFile.cell(row=row, column=6).value = None
        
        tableLog.cell(row=newRow, column=1).value = fileID
        tableLog.cell(row=newRow, column=2).value = folderPath
        tableLog.cell(row=newRow, column=3).value = newFileName
        tableLog.cell(row=newRow, column=4).value = fileName

def main():

    parser = argparse.ArgumentParser(
        description="For maintain filename with spreadsheet."
    )

    parser.add_argument("-r", "--rename", action="store_true", help="Renaming all selected files.")
    parser.add_argument("-i", "--inspect", action="store_true", help="Only update sheet with filename changes.")

    args = parser.parse_args()

    if args.rename:

        # rename file
        rRow, rFileID, oldFileName, folderPath, newFileName = listRename("OK")
        if rRow == []:
            pass
        else:
            for i, j, k, l, m in zip(rRow, rFileID, oldFileName, folderPath, newFileName):
                old = "\\".join([str(l), str(k)])
                a, b = os.path.splitext(old)
                new = "\\".join([str(l), "".join([str(m), b.upper()])])
                os.rename(old, new)

                insertData(
                    status="Renamed",
                    row=i,
                    folderPath=l,
                    fileID=j,
                    newFileName="".join([str(m), b.upper()]),
                    fileName=k,
                    newRow=lastRow(tableLog["A"])+1
                )
    else:
        pass

    if args.inspect:

        for x in listFolder:

            fileID, fileName = listFile(x)
            dataRow, dataFileID, dataFileName = listData(x)
            
            # check existing data list
            for i, j, k in zip(dataRow, dataFileID, dataFileName):
                if j in fileID and k in fileName:
                    insertData(
                        status="Existing",
                        row=i,
                        folderPath=x
                    )
                elif j not in fileID and k not in fileName:
                    insertData(
                        status="Missing",
                        row=i,
                        folderPath=x
                    )

            # check for new file
            for i, j in zip(fileID, fileName):
                if i not in dataFileID and j not in dataFileName:
                    insertData(
                        status="New",
                        row=lastRow(tableFile["A"])+1,
                        folderPath=x,
                        fileID=i,
                        fileName=j
                    )
    else:
        pass

    try:
        xlsx.save(xlsxFile)
    except:
        print("The LOG.XLSX is still running, please close the file first and run the script again.")
    
if __name__ == "__main__":
    main()