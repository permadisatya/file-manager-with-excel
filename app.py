#!/usr/bin/env python

from genericpath import exists
import os
import pathlib
from openpyxl import load_workbook, Workbook
from posixpath import join

fpath = r"C:\Users\permadisatya\Documents\DT Documents\Personal\Log"
txt_filename = "FOLDER.TXT"
xlsx_filename = os.path.join(fpath, "LOG.XLSX")
txt = os.path.join(fpath, txt_filename)
xlsx = load_workbook(xlsx_filename)


sheet = xlsx["tbl_update"]
log = xlsx["tbl_log"]

# function to read all file information in certain folder
def listFile(pathfolder):
    a = []
    b = []
    for entry in pathlib.Path(pathfolder).iterdir():
        if entry.is_file():
            a.append("".join(["ID", str(os.path.getctime(entry)).replace(".", "")]))
            b.append(entry.name)
    return a, b

# function to update sheet
def insertData(status, row, pathfolder, id = None, filename = None, newfilename = None, oldfilename = None, row_log = None):
    
    sheet.cell(row = row, column = 3).value = pathfolder
    sheet.cell(row = row, column = 4).value = status

    if status == "New":
        sheet.cell(row = row, column = 1).value = id
        sheet.cell(row = row, column = 2).value = filename

    elif status == "Renamed":
        sheet.cell(row = row, column = 2).value = newfilename
        log.cell(row = row_log, column = 1).value = id
        log.cell(row = row_log, column = 2).value = pathfolder
        log.cell(row = row_log, column = 3).value = newfilename
        log.cell(row = row_log, column = 4).value = oldfilename

# function to get last non-empty row
def lastRow(col):
    row = []
    for i in col:
        if i.value != "" and i.value != None:
            row.append(i.row)
    return max(row)

def main():
    file = open(txt).read().splitlines()
    folder = []
    for p in file:
        folder.append(p.lstrip())
    for fp in folder:

        # check status existing excel sheet
        file_id, file_name = listFile(fp)
        for x, y, z in zip(sheet["A"], sheet["B"], sheet["C"]):
            if x.value != "id_file" and y.value != "current_file_name":
                if z.value != fp:
                    pass
                else:
                    if x.value in file_id and y.value in file_name:
                        a = x.row
                        insertData(status = "Existing", row = a, pathfolder = fp)

                    elif x.value in file_id and y.value not in file_name:
                        a = x.row
                        b = x.value
                        c = lastRow(log["A"]) + 1
                        d = sheet.cell(row = x.row, column = 2).value
                        e = file_name[file_id.index(x.value)]
                        insertData(
                            status = "Renamed", 
                            row = a, 
                            pathfolder = fp, 
                            id = b, 
                            newfilename = e, 
                            oldfilename = d, 
                            row_log = c
                        )

                    elif x.value not in file_id:
                        a = x.row
                        insertData(status = "Missing", row = a, pathfolder = fp)

        # check file in excel sheet
        for id in file_id:
            if id not in [c.value for c in sheet["A"]]:
                a = file_name[file_id.index(id)]
                b = "New"              
                c = lastRow(sheet["A"]) + 1
                
                insertData(id = id, filename = a, status = b, row  = c, pathfolder = fp)

        # bulk rename
        for x in sheet["G"]:
            if x.value == "ok":

                a = x.row
                b = sheet.cell(row = a, column = 1).value
                c = file_name[file_id.index(b)]
                d = "\\".join([str(folder), str(c)])
                e, f = os.path.splitext(d)
                g = "\\".join([str(folder), "".join([str(sheet.cell(row = a, column = 6).value), f.upper()])])

                os.rename(d, g)
                
                sheet.cell(row = x.row, column = 6).value = None
                sheet.cell(row = x.row, column = 7).value = None
    
    try:
        xlsx.save(xlsx_filename)
    except:
        print("The LOG.XLSX is still running, please close the file first and run the script again.")
    
if __name__ == "__main__":
    main()