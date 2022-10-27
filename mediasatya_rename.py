import os
import pathlib
from openpyxl import load_workbook, Workbook
from posixpath import join

folder = r"C:\Users\permadisatya\Downloads\Trash"
t = folder.split("\\")
t = t[len(t) - 1]
path_wb = r"C:\Users\permadisatya\Documents\DT Documents\Personal\Log\LOG.XLSX"
wb = load_workbook(path_wb)
sheet = wb["tbl_update"]

file_id = []
file_name = []

for entry in pathlib.Path(folder).iterdir():
    if entry.is_file():
        file_name.append(entry.name)
        file_id.append("".join(["ID", str(os.path.getctime(entry)).replace(".", "")]))

def main():
    # check status existing excel sheet
    for x, y in zip(sheet["A"], sheet["B"]):
        if x.value != "id_file" and y.value != "current_file_name":
            if x.value in file_id and y.value in file_name:
                sheet.cell(row = x.row, column = 4).value = "Exist"
                sheet.cell(row = x.row, column = 3).value = t
            elif x.value in file_id and y.value not in file_name:
                sheet.cell(row = x.row, column = 4).value = "Renamed"
                sheet.cell(row = x.row, column = 5).value = sheet.cell(row = x.row, column = 2).value
                sheet.cell(row = x.row, column = 2).value = file_name[file_id.index(x.value)]
                sheet.cell(row = x.row, column = 3).value = t
            elif x.value not in file_id:
                sheet.cell(row = x.row, column = 4).value = "Missing"
                sheet.cell(row = x.row, column = 3).value = t

    # check file in excel sheet
    for id in file_id:
        if id not in [c.value for c in sheet["A"]]:
            row = []
            for i in sheet["A"]:
                if i.value != "" and i.value != None:
                    row.append(i.row)
            sheet.cell(row = max(row)+1, column = 1).value = id
            sheet.cell(row = max(row)+1, column = 2).value = file_name[file_id.index(id)]
            sheet.cell(row = max(row)+1, column = 3).value = t
            sheet.cell(row = max(row)+1, column = 4).value = "New"

    # bulk rename
    for x in sheet["G"]:
        if x.value == "ok":
            id = sheet.cell(row = x.row, column = 1).value
            name = file_name[file_id.index(id)]
            src = "\\".join([str(folder), str(name)])
            pnm, ext = os.path.splitext(src)
            new = "\\".join([str(folder), "".join([str(sheet.cell(row = x.row, column = 6).value), ext.upper()])])
            os.rename(src, new)
            sheet.cell(row = x.row, column = 6).value = None
            sheet.cell(row = x.row, column = 7).value = None

    wb.save(path_wb)

if __name__ == "__main__":
    main()